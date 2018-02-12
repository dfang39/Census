import openpyxl
import re
import os


alaskaFile = 'C:\\Users\\DanFang\\Desktop\\Alaska Counties.xlsx'
alaskaWB = openpyxl.load_workbook(alaskaFile, data_only = True, read_only = True)
alaskaSheet = alaskaWB.get_sheet_by_name('Sheet1')

alaska = ""
for row in range(1, alaskaSheet.max_row + 1):
    alaska += ',' + str(alaskaSheet['A' + str(row)].value)


print(alaska)

countyFile = 'C:\\Users\\DanFang\\Desktop\\GoogleCountiesPython.xlsx'

countyWB = openpyxl.load_workbook(countyFile, data_only = True, read_only = True)
sheet = countyWB.get_sheet_by_name('Sheet1')


class County(object):
    def __init__(self, code = "", name = "", state = "", variables = "", polygons = [], center = ""):
        self.code = code
        self.name = name
        self.state = state
        self.variables = variables
        self.polygons = polygons
        self.center = center



counties = []


for row in range(2, sheet.max_row + 1):#
    county = County()
    raw_code = str(sheet['H' + str(row)].value)
    if len(raw_code) == 4:
        raw_code = "0" + raw_code
    county.code = raw_code

    county.name = str(sheet['A' + str(row)].value)

    county.state = str(sheet['D' + str(row)].value)

    regex = r"\[[^\]]*\]"
    raw_polygon = str(sheet['E' + str(row)].value)
    polygon_list = re.findall(regex, raw_polygon)


    print(county.code)

    lng_regex = r"lng:[^,]*"
    lat_regex = r"lat:[^}]*"


    raw_longitudes = re.findall(lng_regex, raw_polygon)
    raw_latitudes = re.findall(lat_regex, raw_polygon)


    longitudes = [float(l.replace("lng:", "")) for l in raw_longitudes]
    latitudes = [float(l.replace("lat:", "")) for l in raw_latitudes]


    county.center = "{lat:" + str(sum(latitudes)/len(latitudes)) + ",lng:" + str(sum(longitudes)/len(longitudes)) + "}"

    variables_string = "";
    polygon_string = "";

    for i in range(0, len(polygon_list)):

        variables_string += "coords" + str(i) + ","
        polygon_numbered = "var coords" + str(i) + " = " + polygon_list[i] + "; "
        polygon_string += polygon_numbered


    variables_string_stripped = variables_string[0:-1]
    print(polygon_string)
    print(variables_string_stripped)

    county.polygons = polygon_string
    county.variables = variables_string_stripped

    counties.append(county)


out_folder = "C:\\Users\\DanFang\\Desktop"
os.chdir(out_folder)
out_workbook = openpyxl.Workbook()

sheet = out_workbook.active

row_count = 1
for county in counties:
    sheet['A' + str(row_count)] = county.code
    sheet['B' + str(row_count)] = county.name
    sheet['C' + str(row_count)] = county.state
    sheet['D' + str(row_count)] = county.center
    sheet['E' + str(row_count)] = county.variables
    sheet['F' + str(row_count)] = county.polygons

    row_count += 1


out_workbook.save("CountiesPolygonImport.xlsx")